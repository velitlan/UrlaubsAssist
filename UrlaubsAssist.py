import tkinter as tk
from tkinter import simpledialog, messagebox
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

DATEI = "urlaube.json"
STANDARD_URLAUBSTAGE = 30

def initialisiere_datei():
    if not os.path.exists(DATEI) or os.stat(DATEI).st_size == 0:
        speichere_daten({"mitarbeiter": {}, "antraege": []})

def lade_daten():
    try:
        with open(DATEI, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        speichere_daten({"mitarbeiter": {}, "antraege": []})
        return {"mitarbeiter": {}, "antraege": []}

def speichere_daten(daten):
    with open(DATEI, "w", encoding="utf-8") as f:
        json.dump(daten, f, indent=2)

def mitarbeiter_hinzufuegen_gui():
    name = simpledialog.askstring("Mitarbeiter hinzufügen", "Name des Mitarbeiters:")
    if not name:
        return
    daten = lade_daten()
    if name in daten["mitarbeiter"]:
        messagebox.showwarning("Fehler", "Mitarbeiter existiert bereits.")
        return
    eingabe = simpledialog.askstring("Urlaubstage", f"Urlaubstage für {name} (leer für Standard = {STANDARD_URLAUBSTAGE}):")
    try:
        tage = int(eingabe) if eingabe else STANDARD_URLAUBSTAGE
        if tage <= 0:
            raise ValueError
    except:
        messagebox.showerror("Fehler", "Ungültige Zahl für Urlaubstage.")
        return
    daten["mitarbeiter"][name] = {"urlaubstage": tage}
    speichere_daten(daten)
    messagebox.showinfo("Erfolg", f"{name} mit {tage} Urlaubstagen hinzugefügt.")

def urlaub_beantragen_gui():
    name = simpledialog.askstring("Urlaubsantrag", "Name des Mitarbeiters:")
    if not name:
        return
    try:
        tage = int(simpledialog.askstring("Tage", "Anzahl der Urlaubstage:"))
        if tage <= 0:
            raise ValueError
    except:
        messagebox.showerror("Fehler", "Bitte eine gültige positive Zahl eingeben.")
        return
    daten = lade_daten()
    if name not in daten["mitarbeiter"]:
        messagebox.showwarning("Fehler", "Mitarbeiter nicht gefunden.")
        return
    daten["antraege"].append({"name": name, "tage": tage, "status": "offen"})
    speichere_daten(daten)
    messagebox.showinfo("Gespeichert", "Urlaubsantrag gespeichert.")

def antraege_anzeigen_gui():
    daten = lade_daten()
    antraege = daten.get("antraege", [])
    if not antraege:
        messagebox.showinfo("Info", "Keine Anträge vorhanden.")
        return
    text = "\n".join(f"{i}: {a['name']} - {a['tage']} Tage ({a['status']})" for i, a in enumerate(antraege))
    messagebox.showinfo("Anträge", text)

def antrag_bearbeiten_gui():
    daten = lade_daten()
    antraege = daten["antraege"]
    if not antraege:
        messagebox.showinfo("Info", "Keine Anträge zu bearbeiten.")
        return
    index = simpledialog.askinteger("Bearbeiten", f"Antragsnummer (0 - {len(antraege)-1}):")
    if index is None or not (0 <= index < len(antraege)):
        messagebox.showerror("Fehler", "Ungültige Nummer.")
        return
    antrag = antraege[index]
    entscheidung = messagebox.askquestion("Bearbeiten", f"{antrag['name']} - {antrag['tage']} Tage\nGenehmigen?")
    if entscheidung == "yes":
        name = antrag["name"]
        tage = antrag["tage"]
        rest = daten["mitarbeiter"][name]["urlaubstage"]
        if rest >= tage:
            daten["mitarbeiter"][name]["urlaubstage"] -= tage
            antrag["status"] = "genehmigt"
            messagebox.showinfo("Genehmigt", "Antrag genehmigt.")
        else:
            messagebox.showwarning("Fehler", f"Nicht genügend Urlaubstage (verfügbar: {rest}).")
    else:
        antrag["status"] = "abgelehnt"
        messagebox.showinfo("Abgelehnt", "Antrag abgelehnt.")
    speichere_daten(daten)

def exportiere_excel_daten():
    daten = lade_daten()
    wb = Workbook()
    ws_mitarbeiter = wb.active
    ws_mitarbeiter.title = "Mitarbeiter"
    ws_mitarbeiter.append(["Name", "Verfügbare Urlaubstage", "Genommene Urlaubstage", "Verbleibende Urlaubstage"])
    genommene_tage = {}
    for antrag in daten["antraege"]:
        if antrag["status"] == "genehmigt":
            name = antrag["name"]
            tage = antrag["tage"]
            genommene_tage[name] = genommene_tage.get(name, 0) + tage
    for name, info in daten["mitarbeiter"].items():
        genommen = genommene_tage.get(name, 0)
        ws_mitarbeiter.append([name, info["urlaubstage"] + genommen, genommen, info["urlaubstage"]])
    spaltenbreiten = [25, 25, 25, 35]
    for i, breite in enumerate(spaltenbreiten, start=1):
        ws_mitarbeiter.column_dimensions[get_column_letter(i)].width = breite
    for row in ws_mitarbeiter.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for cell in ws_mitarbeiter["1:1"]:
        cell.font = Font(bold=True)
    ws_antraege = wb.create_sheet(title="Anträge")
    ws_antraege.append(["Name", "Tage", "Status"])
    for antrag in daten["antraege"]:
        ws_antraege.append([antrag["name"], antrag["tage"], antrag["status"]])
    for col in range(1, 4):
        ws_antraege.column_dimensions[get_column_letter(col)].width = 30
    for row in ws_antraege.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for cell in ws_antraege["1:1"]:
        cell.font = Font(bold=True)
    wb.save("urlaubsverwaltung.xlsx")
    messagebox.showinfo("Erfolg", "Excel-Datei erstellt: urlaubsverwaltung.xlsx")

def starte_gui():
    root = tk.Tk()
    root.title("Urlaubsverwaltung")
    root.geometry("300x400")

    tk.Label(root, text="Urlaubsverwaltung", font=("Arial", 16, "bold")).pack(pady=20)

    tk.Button(root, text="Mitarbeiter hinzufügen", width=25, command=mitarbeiter_hinzufuegen_gui).pack(pady=5)
    tk.Button(root, text="Urlaubsantrag stellen", width=25, command=urlaub_beantragen_gui).pack(pady=5)
    tk.Button(root, text="Anträge anzeigen", width=25, command=antraege_anzeigen_gui).pack(pady=5)
    tk.Button(root, text="Antrag bearbeiten", width=25, command=antrag_bearbeiten_gui).pack(pady=5)
    tk.Button(root, text="Excel exportieren", width=25, command=exportiere_excel_daten).pack(pady=5)
    tk.Button(root, text="Beenden", width=25, command=root.destroy).pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    initialisiere_datei()
    starte_gui()